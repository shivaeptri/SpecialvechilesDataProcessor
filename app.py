import io
import zipfile
from datetime import datetime

import pandas as pd
import streamlit as st
import matplotlib.pyplot as plt
import matplotlib.transforms as mtransforms

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment


# ---------------------------------
# STREAMLIT PAGE CONFIG
# ---------------------------------
st.set_page_config(
    page_title="Special Vehicles Reporting Processor",
    page_icon="📊",
    layout="wide"
)

st.title("📊 Special Vehicles Reporting Processor")
st.caption("Upload the Excel file, clean the sheets, generate summaries, charts, and download the outputs.")


# ---------------------------------
# NORMALIZATION FUNCTIONS
# ---------------------------------
def normalize_zone_name(zone):
    if not isinstance(zone, str):
        return zone

    z = zone.strip().lower().replace(" ", "")
    typo_map = {
        "serilinampally": "Serilingampally",
        "serilingampally": "Serilingampally",
        "selingampally": "Serilingampally",
        "sereelingampally": "Serilingampally",
        "srelingampally": "Serilingampally",
        "sreelingampally": "Serilingampally",
        "khaitabad": "Khairatabad",
        "khairatabad": "Khairatabad",
        "golconda": "Golkonda",
        "golkonda": "Golkonda",
    }
    return typo_map.get(z, zone.strip().title())


def normalize_corp_name(corp):
    if not isinstance(corp, str):
        return corp

    c = corp.strip().lower().replace("-", "").replace(" ", "")
    corp_typo_map = {
        "cmc": "CMC",
        "ghmc": "GHMC",
        "mmc": "MMC",
        "cmc-": "CMC",
        "ghmc-": "GHMC",
        "mmc-": "MMC",
    }
    return corp_typo_map.get(c, corp.strip().upper())


# ---------------------------------
# EXCEL PROCESSING FUNCTION
# ---------------------------------
def process_excel(uploaded_file):
    final_columns = [
        "Sno", "date", "Corporation", "Zone", "Circle", "Vehicles Deployed", "LE Deployed"
    ]

    xls = pd.ExcelFile(uploaded_file)
    sheet_names = xls.sheet_names

    cleaned_data = {}
    concat_list = []

    for sheet in sheet_names:
        datafile = pd.read_excel(uploaded_file, sheet_name=sheet)

        if datafile.empty or len(datafile) < 4:
            continue

        try:
            datafile_clean = datafile.iloc[2:].reset_index(drop=True)
            datafile_clean.columns = datafile.iloc[2].values
            datafile_clean = datafile_clean[1:].reset_index(drop=True)

            if len(datafile_clean) > 0:
                datafile_clean = datafile_clean[:-1]

            datafile_clean.columns = [str(col).strip() for col in datafile_clean.columns]
            datafile_clean.insert(1, "date", sheet)

            if "Zone" in datafile_clean.columns:
                datafile_clean["Zone"] = datafile_clean["Zone"].apply(normalize_zone_name)

            if "Corporation" in datafile_clean.columns:
                datafile_clean["Corporation"] = datafile_clean["Corporation"].apply(normalize_corp_name)

            concat_list.append(datafile_clean.copy())

            cleaned_cols = [col for col in final_columns if col != "Sno" and col in datafile_clean.columns]
            individual_cleaned = datafile_clean[cleaned_cols]
            cleaned_data[sheet] = individual_cleaned.copy()

        except Exception:
            continue

    if not concat_list:
        raise ValueError("No valid sheets were found in the uploaded workbook.")

    concatenated_df = pd.concat(concat_list, ignore_index=True)

    # Drop Sno if present
    for col in concatenated_df.columns:
        if str(col).strip().lower() == "sno":
            concatenated_df = concatenated_df.drop(columns=[col])
            break

    for num_col in ["Vehicles Deployed", "LE Deployed"]:
        if num_col in concatenated_df.columns:
            concatenated_df[num_col] = pd.to_numeric(
                concatenated_df[num_col], errors="coerce"
            ).fillna(0).astype(int)

    expected_cols = [col for col in final_columns if col != "Sno" and col in concatenated_df.columns]
    ordered_df = concatenated_df[expected_cols].copy()
    ordered_df.insert(0, "Sno", range(1, len(ordered_df) + 1))

    # Corporation Summary
    corp_sum_df = (
        concatenated_df.groupby("Corporation", as_index=False)[["Vehicles Deployed", "LE Deployed"]]
        .sum()
        .loc[:, ["Corporation", "Vehicles Deployed", "LE Deployed"]]
    )

    # Zone Summary
    zone_sum_df = (
        concatenated_df.groupby(["Corporation", "Zone"], as_index=False)[["Vehicles Deployed", "LE Deployed"]]
        .sum()
        .loc[:, ["Corporation", "Zone", "Vehicles Deployed", "LE Deployed"]]
        .sort_values(["Corporation", "Zone"])
        .reset_index(drop=True)
    )

    zone_sum_df["Corporation - Zone"] = (
        zone_sum_df["Corporation"].astype(str) + " - " + zone_sum_df["Zone"].astype(str)
    )
    zone_sum_df = zone_sum_df[
        ["Corporation - Zone", "Corporation", "Zone", "Vehicles Deployed", "LE Deployed"]
    ]

    # ---------------------------------
    # CLEANED WORKBOOK
    # ---------------------------------
    cleaned_buffer = io.BytesIO()
    with pd.ExcelWriter(cleaned_buffer, engine="openpyxl") as writer:
        for sheet, df in cleaned_data.items():
            safe_sheet = str(sheet)[:31]
            df.to_excel(writer, sheet_name=safe_sheet, index=False)
    cleaned_buffer.seek(0)

    # ---------------------------------
    # CONCATENATED WORKBOOK
    # ---------------------------------
    concat_buffer = io.BytesIO()
    with pd.ExcelWriter(concat_buffer, engine="openpyxl") as writer:
        ordered_df.to_excel(writer, sheet_name="AllData", index=False)
        corp_sum_df.to_excel(writer, sheet_name="Corporation Summary", index=False)
        zone_sum_df.to_excel(writer, sheet_name="Zone Summary", index=False)
    concat_buffer.seek(0)

    wb = load_workbook(concat_buffer)

    # ---------------------------------
    # MERGE CORPORATION COLUMN IN ALLDATA
    # ---------------------------------
    ws_all = wb["AllData"]

    corp_col_idx = None
    for idx, col in enumerate(ws_all[1], 1):
        if col.value == "Corporation":
            corp_col_idx = idx
            break

    if corp_col_idx and ws_all.max_row >= 2:
        start_row = 2
        last_val = ws_all.cell(row=start_row, column=corp_col_idx).value
        merge_start = start_row

        for row in range(start_row + 1, ws_all.max_row + 1):
            curr_val = ws_all.cell(row=row, column=corp_col_idx).value
            if curr_val != last_val:
                if row - 1 > merge_start:
                    ws_all.merge_cells(
                        start_row=merge_start,
                        start_column=corp_col_idx,
                        end_row=row - 1,
                        end_column=corp_col_idx
                    )
                    ws_all.cell(row=merge_start, column=corp_col_idx).alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                merge_start = row
                last_val = curr_val

        if ws_all.max_row > merge_start:
            ws_all.merge_cells(
                start_row=merge_start,
                start_column=corp_col_idx,
                end_row=ws_all.max_row,
                end_column=corp_col_idx
            )
            ws_all.cell(row=merge_start, column=corp_col_idx).alignment = Alignment(
                horizontal="center", vertical="center"
            )

    # ---------------------------------
    # MERGE CORPORATION COLUMN IN ZONE SUMMARY
    # ---------------------------------
    ws_zone_tbl = wb["Zone Summary"]
    zone_tbl_end = 1 + len(zone_sum_df)
    corp_col_b = 2

    if zone_tbl_end >= 2:
        def cell_key(v):
            if v is None:
                return ""
            return str(v).strip().upper()

        ms = 2
        lv = ws_zone_tbl.cell(row=2, column=corp_col_b).value

        for row in range(3, zone_tbl_end + 1):
            cv = ws_zone_tbl.cell(row=row, column=corp_col_b).value
            if cell_key(cv) != cell_key(lv):
                if row - 1 > ms:
                    ws_zone_tbl.merge_cells(
                        start_row=ms,
                        start_column=corp_col_b,
                        end_row=row - 1,
                        end_column=corp_col_b
                    )
                    ws_zone_tbl.cell(row=ms, column=corp_col_b).alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                ms = row
                lv = cv

        if zone_tbl_end > ms:
            ws_zone_tbl.merge_cells(
                start_row=ms,
                start_column=corp_col_b,
                end_row=zone_tbl_end,
                end_column=corp_col_b
            )
            ws_zone_tbl.cell(row=ms, column=corp_col_b).alignment = Alignment(
                horizontal="center", vertical="center"
            )

    # ---------------------------------
    # EXCEL CHARTS
    # ---------------------------------
    chart_blue = "4472C4"
    chart_orange = "ED7D31"

    # Corporation Summary Chart
    ws_corp = wb["Corporation Summary"]
    corp_rows = ws_corp.max_row

    if corp_rows >= 2:
        corp_chart = BarChart()
        corp_chart.title = "Vehicles and LE Deployed by Corporation"
        corp_chart.y_axis.title = "Count"
        corp_chart.width = 16
        corp_chart.height = 7

        data_corp = Reference(ws_corp, min_col=2, min_row=1, max_col=3, max_row=corp_rows)
        cats_corp = Reference(ws_corp, min_col=1, min_row=2, max_row=corp_rows)
        corp_chart.add_data(data_corp, titles_from_data=True)
        corp_chart.set_categories(cats_corp)

        if len(corp_chart.series) > 0:
            corp_chart.series[0].graphicalProperties.solidFill = chart_blue
        if len(corp_chart.series) > 1:
            corp_chart.series[1].graphicalProperties.solidFill = chart_orange

        corp_chart.dLbls = DataLabelList()
        corp_chart.dLbls.showVal = True
        corp_chart.dLbls.showLegendKey = False
        corp_chart.dLbls.showSerName = False
        corp_chart.dLbls.showCatName = False
        corp_chart.dLbls.position = "outEnd"

        ws_corp.add_chart(corp_chart, "E2")

    # Zone Summary Chart with multi-level axis in Excel
    ws_zone = wb["Zone Summary"]
    zone_rows = 1 + len(zone_sum_df)

    if zone_rows >= 2:
        zone_chart = BarChart()
        zone_chart.grouping = "clustered"
        zone_chart.title = "Vehicles and LE Deployed by Corporation & Zone"
        zone_chart.y_axis.title = "Count"
        zone_chart.width = 32
        zone_chart.height = 11

        data_zone = Reference(ws_zone, min_col=4, min_row=1, max_col=5, max_row=zone_rows)
        cats_zone = Reference(ws_zone, min_col=2, min_row=2, max_col=3, max_row=zone_rows)
        zone_chart.add_data(data_zone, titles_from_data=True)
        zone_chart.set_categories(cats_zone)

        if len(zone_chart.series) > 0:
            zone_chart.series[0].graphicalProperties.solidFill = chart_blue
        if len(zone_chart.series) > 1:
            zone_chart.series[1].graphicalProperties.solidFill = chart_orange

        zone_chart.dLbls = DataLabelList()
        zone_chart.dLbls.showVal = True
        zone_chart.dLbls.showLegendKey = False
        zone_chart.dLbls.showSerName = False
        zone_chart.dLbls.showCatName = False
        zone_chart.dLbls.position = "outEnd"

        ws_zone.add_chart(zone_chart, "G2")

    final_concat_buffer = io.BytesIO()
    wb.save(final_concat_buffer)
    final_concat_buffer.seek(0)

    return {
        "cleaned_buffer": cleaned_buffer,
        "concat_buffer": final_concat_buffer,
        "ordered_df": ordered_df,
        "corp_sum_df": corp_sum_df,
        "zone_sum_df": zone_sum_df,
        "sheet_count": len(cleaned_data),
        "row_count": len(ordered_df),
        "vehicle_total": int(ordered_df["Vehicles Deployed"].sum()) if "Vehicles Deployed" in ordered_df.columns else 0,
        "le_total": int(ordered_df["LE Deployed"].sum()) if "LE Deployed" in ordered_df.columns else 0,
    }


# ---------------------------------
# MATPLOTLIB CHART HELPERS
# ---------------------------------
def plot_corporation_summary(corp_sum_df):
    fig, ax = plt.subplots(figsize=(10, 5))

    x = range(len(corp_sum_df))
    width = 0.35

    bars1 = ax.bar(
        [i - width / 2 for i in x],
        corp_sum_df["Vehicles Deployed"],
        width,
        label="Vehicles Deployed"
    )
    bars2 = ax.bar(
        [i + width / 2 for i in x],
        corp_sum_df["LE Deployed"],
        width,
        label="LE Deployed"
    )

    ax.set_xticks(list(x))
    ax.set_xticklabels(corp_sum_df["Corporation"], rotation=0)
    ax.set_title("Vehicles and LE Deployed by Corporation")
    ax.set_ylabel("Count")
    ax.legend()

    for bars in [bars1, bars2]:
        for b in bars:
            h = b.get_height()
            ax.text(
                b.get_x() + b.get_width() / 2,
                h + 1,
                f"{int(h)}",
                ha="center",
                va="bottom",
                fontsize=9
            )

    fig.tight_layout()
    return fig


def plot_zone_summary_multilevel(zone_sum_df):
    # Multi-level x-axis:
    # inner labels = Zone
    # outer grouped labels = Corporation

    if zone_sum_df.empty:
        fig, ax = plt.subplots()
        ax.set_title("No data available")
        return fig

    width = 0.35
    gap_within_group = 0.18
    gap_between_corporations = 0.90

    veh_x = []
    le_x = []
    tick_x = []
    tick_labels = []
    corp_centers = []

    x_cursor = 0.0

    grouped = zone_sum_df.groupby("Corporation", sort=False)

    for corp, grp in grouped:
        grp = grp.reset_index(drop=True)
        n = len(grp)
        first_center = None
        last_center = None

        for i in range(n):
            center = x_cursor
            if first_center is None:
                first_center = center
            last_center = center

            veh_x.append(center - width / 2)
            le_x.append(center + width / 2)
            tick_x.append(center)
            tick_labels.append(str(grp.loc[i, "Zone"]))

            x_cursor += (2 * width + gap_within_group)

        corp_centers.append((corp, (first_center + last_center) / 2))
        x_cursor += gap_between_corporations

    fig_width = max(12, len(zone_sum_df) * 1.0)
    fig, ax = plt.subplots(figsize=(fig_width, 6.5))

    bars1 = ax.bar(veh_x, zone_sum_df["Vehicles Deployed"], width, label="Vehicles Deployed")
    bars2 = ax.bar(le_x, zone_sum_df["LE Deployed"], width, label="LE Deployed")

    # Inner axis labels = Zone
    ax.set_xticks(tick_x)
    ax.set_xticklabels(tick_labels, rotation=45, ha="right")

    # Outer grouped labels = Corporation
    trans = mtransforms.blended_transform_factory(ax.transData, ax.transAxes)
    for corp, cx in corp_centers:
        ax.text(
            cx,
            -0.28,
            corp,
            transform=trans,
            ha="center",
            va="top",
            fontsize=10,
            fontweight="bold"
        )

    # Optional vertical separators between corporation groups
    prev_end = None
    grouped2 = zone_sum_df.groupby("Corporation", sort=False)
    running_idx = 0
    for corp, grp in grouped2:
        grp_len = len(grp)
        if running_idx > 0:
            boundary_index = running_idx - 0.5
            boundary_x = tick_x[boundary_index] if isinstance(boundary_index, int) else None
        running_idx += grp_len

    ax.set_ylabel("Count")
    ax.set_title("Vehicles and LE Deployed by Corporation & Zone")
    ax.legend()

    # Data labels on bars
    for bars in [bars1, bars2]:
        for b in bars:
            h = b.get_height()
            ax.text(
                b.get_x() + b.get_width() / 2,
                h + 1,
                f"{int(h)}",
                ha="center",
                va="bottom",
                fontsize=8
            )

    fig.subplots_adjust(bottom=0.32)
    fig.tight_layout()
    return fig


# ---------------------------------
# SIDEBAR
# ---------------------------------
with st.sidebar:
    st.header("Upload Excel File")
    uploaded_file = st.file_uploader(
        "Choose an Excel workbook",
        type=["xlsx", "xlsm", "xls"]
    )

    process_btn = st.button("🚀 Process File", use_container_width=True)

    st.markdown("---")
    st.subheader("Outputs")
    st.write("• Cleaned workbook")
    st.write("• Final workbook")
    st.write("• Corporation summary")
    st.write("• Zone summary")
    st.write("• Embedded Excel charts")


# ---------------------------------
# MAIN APP
# ---------------------------------
if uploaded_file is None:
    st.info("Please upload an Excel workbook to continue.")
else:
    st.success(f"Loaded file: **{uploaded_file.name}**")

    if process_btn:
        with st.spinner("Processing workbook..."):
            try:
                result = process_excel(uploaded_file)

                cleaned_buffer = result["cleaned_buffer"]
                concat_buffer = result["concat_buffer"]
                ordered_df = result["ordered_df"]
                corp_sum_df = result["corp_sum_df"]
                zone_sum_df = result["zone_sum_df"]

                c1, c2, c3, c4 = st.columns(4)
                c1.metric("Valid Sheets Processed", result["sheet_count"])
                c2.metric("Total Rows", result["row_count"])
                c3.metric("Vehicles Deployed", result["vehicle_total"])
                c4.metric("LE Deployed", result["le_total"])

                st.markdown("---")

                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                cleaned_name = f"Cleaned_Special_Vehicles_Reporting_{timestamp}.xlsx"
                concat_name = f"Concatenated_Cleaned_Special_Vehicles_Reporting_{timestamp}.xlsx"

                d1, d2, d3 = st.columns(3)

                with d1:
                    st.download_button(
                        label="⬇️ Download Cleaned Workbook",
                        data=cleaned_buffer.getvalue(),
                        file_name=cleaned_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

                with d2:
                    st.download_button(
                        label="⬇️ Download Final Workbook",
                        data=concat_buffer.getvalue(),
                        file_name=concat_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                    zf.writestr(cleaned_name, cleaned_buffer.getvalue())
                    zf.writestr(concat_name, concat_buffer.getvalue())
                zip_buffer.seek(0)

                with d3:
                    st.download_button(
                        label="📦 Download ZIP",
                        data=zip_buffer.getvalue(),
                        file_name=f"Special_Vehicles_Output_{timestamp}.zip",
                        mime="application/zip",
                        use_container_width=True
                    )

                st.markdown("---")

                tab1, tab2, tab3, tab4 = st.tabs([
                    "All Data Preview",
                    "Corporation Summary",
                    "Zone Summary",
                    "Charts Preview"
                ])

                with tab1:
                    st.subheader("AllData")
                    st.dataframe(ordered_df, use_container_width=True, height=500)

                with tab2:
                    st.subheader("Corporation Summary")
                    st.dataframe(corp_sum_df, use_container_width=True)

                with tab3:
                    st.subheader("Zone Summary")
                    st.dataframe(zone_sum_df, use_container_width=True, height=500)

                with tab4:
                    st.subheader("Chart Preview")

                    st.markdown("### Corporation Summary Chart")
                    fig1 = plot_corporation_summary(corp_sum_df)
                    st.pyplot(fig1, use_container_width=True)

                    st.markdown("### Zone Summary Chart")
                    st.caption("Inner labels = Zone, outer grouped labels = Corporation")
                    fig2 = plot_zone_summary_multilevel(zone_sum_df)
                    st.pyplot(fig2, use_container_width=True)

                st.success("Processing completed successfully.")

            except Exception as e:
                st.error(f"Processing failed: {e}")
